from __future__ import annotations

import csv
import io
import json
from collections.abc import Iterable
from datetime import UTC, datetime
from typing import Any
from urllib.parse import urlencode

from openpyxl import load_workbook

from landintel.config import Settings
from landintel.connectors.base import (
    ConnectorAsset,
    ConnectorContext,
    ConnectorRunOutput,
    ListingConnector,
    ParsedListing,
)
from landintel.data_fetch.http_assets import fetch_http_asset
from landintel.domain.enums import (
    ConnectorType,
    ListingStatus,
    ListingType,
    PriceBasisType,
    SourceParseStatus,
)
from landintel.listings.parsing import build_search_text, normalize_address, normalize_space

DEFAULT_LONDON_AUTHORITY_PATTERNS = [
    "LONDON BOROUGH",
    "ROYAL BOROUGH",
    "CITY OF LONDON",
]
SUPPORTED_TRANSFORMS = {"cabinet_office_surplus_property_v1"}


class GenericTabularFeedConnector(ListingConnector):
    connector_type = ConnectorType.TABULAR_FEED

    def __init__(self, settings: Settings) -> None:
        self.settings = settings

    def run(
        self,
        *,
        context: ConnectorContext,
        payload: dict[str, Any],
    ) -> ConnectorRunOutput:
        del payload

        refresh_policy = context.refresh_policy_json
        feed_url = str(refresh_policy.get("feed_url") or "").strip()
        if not feed_url:
            raise ValueError("Tabular-feed connector requires refresh_policy_json.feed_url")

        row_transform = str(refresh_policy.get("row_transform") or "").strip()
        if row_transform not in SUPPORTED_TRANSFORMS:
            raise ValueError(
                "Tabular-feed connector requires a supported refresh_policy_json.row_transform"
            )

        fetched = fetch_http_asset(
            feed_url,
            timeout_seconds=self.settings.snapshot_http_timeout_seconds,
        )
        rows = _load_rows(
            fetched.content,
            content_type=fetched.content_type,
            refresh_policy=refresh_policy,
        )
        listings = _transform_rows(
            rows=rows,
            refresh_policy=refresh_policy,
            row_transform=row_transform,
            feed_url=fetched.final_url,
            observed_at=fetched.fetched_at,
        )
        max_listings = int(refresh_policy.get("max_listings") or 50)
        listings = listings[:max_listings]

        parse_status = SourceParseStatus.PARSED if listings else SourceParseStatus.FAILED
        asset = ConnectorAsset(
            asset_key="tabular_feed_raw",
            asset_type=_tabular_asset_type(fetched.content_type, fetched.final_url),
            role="TABULAR_FEED",
            original_url=fetched.final_url,
            content=fetched.content,
            content_type=fetched.content_type,
            fetched_at=fetched.fetched_at,
            metadata={
                "status_code": fetched.status_code,
                "headers": fetched.headers,
                "row_count": len(rows),
                "listing_count": len(listings),
                "row_transform": row_transform,
            },
        )
        return ConnectorRunOutput(
            source_name=context.source_name,
            source_family=self.connector_type.value.lower(),
            source_uri=fetched.final_url,
            observed_at=fetched.fetched_at,
            coverage_note=(
                "Automated tabular feed captured London surplus-property opportunities from "
                f"{fetched.final_url}."
            ),
            parse_status=parse_status,
            manifest_json={
                "connector_type": self.connector_type.value,
                "requested_by": context.requested_by,
                "feed_url": fetched.final_url,
                "content_type": fetched.content_type,
                "row_transform": row_transform,
                "row_count": len(rows),
                "listing_count": len(listings),
            },
            assets=[asset],
            listings=listings,
        )


def _load_rows(
    content: bytes,
    *,
    content_type: str,
    refresh_policy: dict[str, Any],
) -> list[dict[str, Any]]:
    feed_format = str(refresh_policy.get("feed_format") or "").strip().lower()
    if not feed_format:
        feed_format = _infer_feed_format(content_type)

    if feed_format == "xlsx":
        return _load_xlsx_rows(
            content,
            sheet_name=str(refresh_policy.get("sheet_name") or "").strip(),
        )
    if feed_format == "csv":
        return _load_csv_rows(content)
    if feed_format == "json":
        return _load_json_rows(content)
    raise ValueError(f"Unsupported tabular feed format: {feed_format}")


def _load_xlsx_rows(content: bytes, *, sheet_name: str) -> list[dict[str, Any]]:
    workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [
        normalize_space(str(value)) or f"column_{index}"
        for index, value in enumerate(rows[0])
    ]
    records: list[dict[str, Any]] = []
    for values in rows[1:]:
        record = {
            headers[index]: values[index] if index < len(values) else None
            for index in range(len(headers))
        }
        if any(value not in (None, "") for value in record.values()):
            records.append(record)
    return records


def _load_csv_rows(content: bytes) -> list[dict[str, Any]]:
    reader = csv.DictReader(io.StringIO(content.decode("utf-8-sig")))
    return [dict(row) for row in reader]


def _load_json_rows(content: bytes) -> list[dict[str, Any]]:
    payload = json.loads(content.decode("utf-8-sig"))
    if isinstance(payload, list):
        return [dict(row) for row in payload if isinstance(row, dict)]
    if isinstance(payload, dict):
        rows = payload.get("rows") or payload.get("items") or payload.get("results") or []
        if isinstance(rows, list):
            return [dict(row) for row in rows if isinstance(row, dict)]
    raise ValueError("Tabular JSON feed did not contain a top-level row list.")


def _transform_rows(
    *,
    rows: Iterable[dict[str, Any]],
    refresh_policy: dict[str, Any],
    row_transform: str,
    feed_url: str,
    observed_at: datetime,
) -> list[ParsedListing]:
    if row_transform == "cabinet_office_surplus_property_v1":
        return _transform_cabinet_office_surplus_property(
            rows=rows,
            refresh_policy=refresh_policy,
            feed_url=feed_url,
            observed_at=observed_at,
        )
    raise ValueError(f"Unsupported row transform: {row_transform}")


def _transform_cabinet_office_surplus_property(
    *,
    rows: Iterable[dict[str, Any]],
    refresh_policy: dict[str, Any],
    feed_url: str,
    observed_at: datetime,
) -> list[ParsedListing]:
    allowed_statuses = {
        value.lower()
        for value in list(
            refresh_policy.get("status_of_sale_values") or ["On the Market", "Under Offer"]
        )
    }
    authority_patterns = [
        value.upper()
        for value in list(
            refresh_policy.get("local_authority_contains_any")
            or DEFAULT_LONDON_AUTHORITY_PATTERNS
        )
    ]
    transformed: list[ParsedListing] = []
    seen_ids: set[str] = set()

    for row in rows:
        disposal_id = _string_value(row.get("In Site Disposal Reference")) or _string_value(
            row.get("Insite Property Reference")
        )
        if not disposal_id or disposal_id in seen_ids:
            continue

        sale_status = _string_value(row.get("Status of Sale")) or "Unknown"
        if sale_status.lower() not in allowed_statuses:
            continue

        if not _is_london_row(row, authority_patterns=authority_patterns):
            continue

        latitude = _float_value(row.get("Latitude"))
        longitude = _float_value(row.get("Longitude"))
        if latitude is None or longitude is None:
            continue

        seen_ids.add(disposal_id)
        headline = _string_value(row.get("Property Name")) or disposal_id
        contract_name = _string_value(row.get("Contract Name"))
        land_usage = _string_value(row.get("Land Usage"))
        local_authority = _string_value(row.get("Local Authority"))
        address = _compose_address(row)
        listing_type = _cabinet_office_listing_type(row)
        if not _cabinet_office_row_allowed(
            row=row,
            listing_type=listing_type,
            refresh_policy=refresh_policy,
        ):
            continue
        description = normalize_space(
            " · ".join(
                part
                for part in [
                    contract_name,
                    land_usage,
                    local_authority,
                    sale_status,
                    _string_value(row.get("Contract Parent Organisation")),
                ]
                if part
            )
        )
        canonical_url = _build_row_url(feed_url, disposal_id)
        transformed.append(
            ParsedListing(
                source_listing_id=disposal_id,
                canonical_url=canonical_url,
                observed_at=observed_at.astimezone(UTC),
                listing_type=listing_type,
                headline=headline,
                description_text=description,
                guide_price_gbp=None,
                price_basis_type=PriceBasisType.UNKNOWN,
                status=_cabinet_office_listing_status(sale_status),
                auction_date=None,
                address_text=address,
                normalized_address=normalize_address(address),
                lat=latitude,
                lon=longitude,
                raw_record_json={key: _json_safe_value(value) for key, value in row.items()},
                search_text=build_search_text(headline, description, address, local_authority),
            )
        )

    return transformed


def _cabinet_office_listing_status(sale_status: str) -> ListingStatus:
    lowered = sale_status.lower()
    if lowered == "under offer":
        return ListingStatus.UNDER_OFFER
    return ListingStatus.LIVE


def _cabinet_office_listing_type(row: dict[str, Any]) -> ListingType:
    land_area = _float_value(row.get("Total Surplus Land Area")) or 0.0
    floor_area = _float_value(row.get("Total Surplus Floor Area")) or 0.0
    land_usage = (_string_value(row.get("Land Usage")) or "").lower()
    text = " ".join(
        part
        for part in [
            _string_value(row.get("Property Name")),
            _string_value(row.get("Contract Name")),
            _string_value(row.get("Land Usage")),
        ]
        if part
    ).lower()

    if land_area > 0 and floor_area <= 0:
        return ListingType.LAND
    if "development" in text or "redevelop" in text:
        return ListingType.REDEVELOPMENT_SITE
    if "surplus land" in land_usage:
        return ListingType.LAND
    return ListingType.LAND_WITH_BUILDING


def _is_london_row(
    row: dict[str, Any],
    *,
    authority_patterns: list[str],
) -> bool:
    local_authority = (_string_value(row.get("Local Authority")) or "").upper()
    if authority_patterns:
        return any(pattern in local_authority for pattern in authority_patterns)
    region = (_string_value(row.get("Region")) or "").upper()
    town = (_string_value(row.get("Town")) or "").upper()
    if "LONDON" in region or "LONDON" in town:
        return True
    return any(pattern in local_authority for pattern in DEFAULT_LONDON_AUTHORITY_PATTERNS)


def _cabinet_office_row_allowed(
    *,
    row: dict[str, Any],
    listing_type: ListingType,
    refresh_policy: dict[str, Any],
) -> bool:
    floor_area = _float_value(row.get("Total Surplus Floor Area")) or 0.0
    max_floor_area = refresh_policy.get("max_surplus_floor_area_sqm")
    if max_floor_area is not None and floor_area > float(max_floor_area):
        return False

    land_area = _float_value(row.get("Total Surplus Land Area")) or 0.0
    if bool(refresh_policy.get("require_positive_land_area")) and land_area <= 0:
        return False

    allowed_land_usage_patterns = [
        str(value).strip().lower()
        for value in list(refresh_policy.get("allowed_land_usage_contains_any") or [])
        if str(value).strip()
    ]
    land_usage = (_string_value(row.get("Land Usage")) or "").lower()
    if allowed_land_usage_patterns and not any(
        pattern in land_usage for pattern in allowed_land_usage_patterns
    ):
        return False

    allowed_listing_types: set[ListingType] = set()
    for value in list(refresh_policy.get("allowed_listing_types") or []):
        normalized = str(value).strip()
        if not normalized:
            continue
        try:
            allowed_listing_types.add(ListingType(normalized))
        except ValueError:
            continue
    return not allowed_listing_types or listing_type in allowed_listing_types


def _compose_address(row: dict[str, Any]) -> str | None:
    address_parts = [
        _string_value(row.get("Property Number")),
        _string_value(row.get("Street Name")),
        _string_value(row.get("Town")),
        _string_value(row.get("Postcode")),
    ]
    return normalize_space(", ".join(part for part in address_parts if part))


def _build_row_url(feed_url: str, disposal_id: str) -> str:
    query = urlencode({"disposal_id": disposal_id})
    return f"{feed_url}{'&' if '?' in feed_url else '?'}{query}"


def _infer_feed_format(content_type: str) -> str:
    lowered = content_type.lower()
    if "spreadsheetml" in lowered or "excel" in lowered or "application/zip" in lowered:
        return "xlsx"
    if "csv" in lowered or "text/plain" in lowered:
        return "csv"
    if "json" in lowered:
        return "json"
    return "xlsx"


def _tabular_asset_type(content_type: str, original_url: str) -> str:
    lowered = content_type.lower()
    if "spreadsheetml" in lowered or original_url.lower().endswith(".xlsx"):
        return "XLSX"
    if "excel" in lowered or original_url.lower().endswith(".xls"):
        return "XLS"
    if "json" in lowered or original_url.lower().endswith(".json"):
        return "JSON"
    return "CSV"


def _string_value(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        return normalize_space(value)
    return normalize_space(str(value))


def _float_value(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _json_safe_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat()
    return value
